import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

# API URL
api_url = "http://192.168.100.131:8080/ords/ultra_inv/ad_ons_api/nvestment"
# File paths
excel_file = 'Comp_inv.xlsx'
receivable_file = 'Comp_inv_receivable.xlsx'


# Function to fetch data from the API
def fetch_data_from_api(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()
        return data
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data from API: {e}")
        return None


# Function to create 'Comp_inv.xlsx' with API data and additional receivable data
def create_excel_with_api_data(api_url, excel_file, receivable_file):
    # Fetch data from the API
    api_data = fetch_data_from_api(api_url)
    if api_data:
        # Convert API data to DataFrame
        api_df = pd.DataFrame(api_data['items'])

        # Print each ID from the API data
        if 'id' in api_df.columns:
            print("IDs from API data:")
            for api_id in api_df['id']:
                print(api_id)

        # Read the receivable Excel file
        try:
            receivable_df = pd.read_excel(receivable_file)
        except FileNotFoundError:
            print(f"Receivable file {receivable_file} not found.")
            return

        # Column mapping to match Excel structure
        column_mapping = {
            'id': 'ID',
            'scheme_id': 'SCHEME_ID',
            'instrument': 'INSTRUMENT',
            'asset_classification': 'ASSET_CLASSIFICATION',
            'issuer_name': 'ISSUER_NAME',
            'fund_manager_who_adviced': 'FUND_MANAGER_WHO_ADVICED',
            'investment_id': 'INVESTMENT_ID',
            'issue_date': 'ISSUE_DATE',
            'asset_class': 'ASSET_CLASS',
            'date_of_investment': 'DATE_OF_INVESTMENT',
            'interest_rate_percent': 'INTEREST_RATE_PERCENT',
            'discount_rate_percent': 'DISCOUNT_RATE_PERCENT',
            'coupon_rate_percent': 'COUPON_RATE_PERCENT',
            'currency_conversion_rate': 'CURRENCY_CONVERSION_RATE',
            'currency': 'CURRENCY',
            'asset_tenor': 'ASSET_TENOR',
            'face_value': 'FACE_VALUE',
            'amount_invested': 'AMOUNT_INVESTED',
            'amount_invested__in_foreign_currency': 'AMOUNT_INVESTED__IN_FOREIGN_CURRENCY',
            'investment_charge_amount': 'INVESTMENT_CHARGE_AMOUNT',
            'price_per_unit_share_at_purchase': 'PRICE_PER_UNIT_SHARE_AT_PURCHASE',
            'price_per_unit_share_at_value_date': 'PRICE_PER_UNIT_SHARE_AT_VALUE_DATE',
            'number_of_units_or_shares': 'NUMBER_OF_UNITS_OR_SHARES'
        }

        # Rename API columns to match Excel
        api_df.rename(columns=column_mapping, inplace=True)

        # Add 'ROWID' column if missing
        if 'ROWID' in receivable_df.columns and 'ROWID' not in api_df.columns:
            api_df['ROWID'] = pd.NA

        # Convert date columns in API data
        date_columns = ['ISSUE_DATE', 'DATE_OF_INVESTMENT']
        for col in date_columns:
            if col in api_df.columns:
                api_df[col] = pd.to_datetime(api_df[col], errors='coerce')

        # Combine receivable_df and api_df
        combined_df = pd.concat([receivable_df, api_df], ignore_index=True)
        combined_df = combined_df.fillna('')

        # Create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Investment Data'

        # Write headers
        headers = combined_df.columns
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_idx, value=header)

        # Write data to sheet
        for row_idx, row_data in combined_df.iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx + 2, column=col_idx)
                if headers[col_idx - 1] in date_columns and pd.notna(value):
                    cell.value = pd.to_datetime(value).date()
                    cell.number_format = 'MM/DD/YY'
                else:
                    cell.value = value

        # Save new Excel file
        workbook.save(excel_file)
        print(f"New Excel file '{excel_file}' created successfully with combined data.")
    else:
        print("No data from API to populate.")


# Run the function
create_excel_with_api_data(api_url, excel_file, receivable_file)
